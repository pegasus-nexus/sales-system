import os
import openpyxl
import pandas as pd

def read_xlsx(path):
    print(f"=== Reading {path} ===")
    if not os.path.exists(path):
        print("File does not exist")
        return
    try:
        # Try pandas first
        df = pd.read_excel(path)
        print(df.to_string())
    except Exception as e:
        print(f"Failed with pandas: {e}")
        try:
            # Try openpyxl
            wb = openpyxl.load_workbook(path)
            for sheet in wb.sheetnames:
                print(f"Sheet: {sheet}")
                ws = wb[sheet]
                for r in list(ws.iter_rows(values_only=True))[:30]:
                    print(r)
        except Exception as e2:
            print(f"Failed with openpyxl: {e2}")

path1 = r"c:\Users\rodri\Desktop\Taboada System\SalesSystem\otros\error de conteo inventario 1.xlsx"
path2 = r"c:\Users\rodri\Desktop\Taboada System\SalesSystem\otros\error de conteo inventario 2.xlsx"

read_xlsx(path1)
read_xlsx(path2)
