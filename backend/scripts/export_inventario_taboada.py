"""
export_inventario_taboada.py
============================
Exporta TODOS los productos e inventario del tenant TABOADA a un Excel.

Genera un archivo Excel con:
  - Hoja "INVENTARIO_COMPLETO"  → una fila por (producto × sucursal × almacen)
  - Hoja "RESUMEN_POR_SUCURSAL" → total de productos y stock por sucursal
  - Hoja "PRODUCTOS_SIN_STOCK"  → productos que existen en catálogo pero tienen stock 0

Uso:
  1. Para base de datos LOCAL:
     cd backend
     python scripts/export_inventario_taboada.py

  2. Para base de datos de PRODUCCIÓN (Atlas):
     $env:MONGODB_URL="mongodb+srv://user:pass@cluster.mongodb.net/dbname"
     python scripts/export_inventario_taboada.py

  El archivo Excel se guardará en el directorio actual: inventario_taboada_YYYY-MM-DD_HH-MM.xlsx
"""

import asyncio
import os
import sys
from datetime import datetime

# Fix Windows console encoding for non-ASCII characters
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# ─── Dependencias ──────────────────────────────────────────────────────────────
try:
    from motor.motor_asyncio import AsyncIOMotorClient
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"\n❌ Dependencia faltante: {e}")
    print("   Instalar con: pip install motor pandas openpyxl")
    sys.exit(1)

# ─── Configuración ─────────────────────────────────────────────────────────────
MONGO_URL = os.environ.get(
    "MONGODB_URL",
    "mongodb://localhost:27017/sales_system"  # Default: DB local
)

OUTPUT_FILE = f"inventario_taboada_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"


def get_db_name_from_url(url: str) -> str:
    """Extrae el nombre de la DB de la connection string."""
    try:
        path = url.split("/")[-1].split("?")[0]
        if path and path not in ["", "27017"]:
            return path
    except Exception:
        pass
    return "sales_system"


def safe_float(val):
    """Convierte Decimal128, str, None → float sin explotar."""
    if val is None:
        return 0.0
    try:
        if hasattr(val, 'to_decimal'):   # Motor Decimal128 de BSON
            return float(val.to_decimal())
        return float(str(val))
    except Exception:
        return 0.0


def format_sheet(ws, header_color="1E3A5F"):
    """Aplica encabezados en color, autofit de columnas y freeze row 1."""
    header_fill = PatternFill("solid", fgColor=header_color)
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max(
            (len(str(cell.value or "")) for cell in col if cell.value),
            default=10
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 48)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


async def export_inventario():
    print(f"\n{'='*62}")
    print("    EXPORTADOR DE INVENTARIO - TENANT TABOADA")
    print(f"{'='*62}")
    url_display = MONGO_URL.split("@")[-1] if "@" in MONGO_URL else MONGO_URL
    print(f"  DB: {url_display}")
    print()

    # ── 1. Conectar ────────────────────────────────────────────────────────────
    client = AsyncIOMotorClient(MONGO_URL, serverSelectionTimeoutMS=6000)
    try:
        await client.admin.command("ping")
        print("  [OK] Conexion exitosa")
    except Exception as e:
        print(f"  [ERROR] No se pudo conectar: {e}")
        print()
        print("  El servidor local no esta corriendo, o la URL no es correcta.")
        print("  Para usar la DB de produccion (Atlas):")
        print('      $env:MONGODB_URL="mongodb+srv://user:pass@cluster.mongodb.net/dbname"')
        print('      python scripts/export_inventario_taboada.py')
        await client.close()
        return

    db_name = get_db_name_from_url(MONGO_URL)
    db = client[db_name]

    # ── 2. Buscar tenant TABOADA ───────────────────────────────────────────────
    print("\n  🔍 Buscando tenant 'TABOADA'...")
    tenant = await db.tenants.find_one(
        {"name": {"$regex": "taboada", "$options": "i"}}
    )

    if not tenant:
        all_tenants = await db.tenants.find({}, {"name": 1}).to_list(length=50)
        print(f"  ❌ Tenant 'TABOADA' no encontrado en la base de datos '{db_name}'.")
        if all_tenants:
            names = [t.get("name", "?") for t in all_tenants]
            print(f"  Tenants disponibles: {', '.join(names)}")
        else:
            print("  ⚠️  No hay tenants registrados en esta base de datos.")
        await client.close()
        return

    tenant_id = str(tenant["_id"])
    tenant_name = tenant.get("name", "TABOADA")
    print(f"  [OK] Tenant: '{tenant_name}'  (ID: {tenant_id})")

    # -- 3. Sucursales ---------------------------------------------------------
    print("\n  [*] Cargando sucursales...")
    sucursales_raw = await db.sucursales.find({"tenant_id": tenant_id}).to_list(length=200)
    sucursales_map = {"CENTRAL": "Matriz / Central"}
    for s in sucursales_raw:
        sucursales_map[str(s["_id"])] = s.get("nombre", "Sin nombre")
    print(f"  [OK] {len(sucursales_raw)} sucursal(es):")
    for sid, nombre in sucursales_map.items():
        print(f"     - {nombre}  ({sid})")

    # -- 4. Categorias ---------------------------------------------------------
    categorias_raw = await db.categories.find({"tenant_id": tenant_id}).to_list(length=1000)
    categorias_map = {str(c["_id"]): c.get("name", "Sin categoria") for c in categorias_raw}
    print(f"\n  [OK] {len(categorias_raw)} categoria(s) encontrada(s)")

    # -- 5. Productos ----------------------------------------------------------
    print("\n  [*] Cargando catalogo de productos...")
    productos_raw = await db.products.find(
        {"tenant_id": tenant_id, "is_active": {"$ne": False}}
    ).to_list(length=50000)
    print(f"  [OK] {len(productos_raw)} producto(s) activo(s)")

    # -- 6. Inventario ---------------------------------------------------------
    print("\n  [*] Cargando registros de inventario...")
    inventario_raw = await db.inventario.find({"tenant_id": tenant_id}).to_list(length=200000)
    print(f"  [OK] {len(inventario_raw)} registro(s) de inventario")

    # ── 7. Indexar inventario ─────────────────────────────────────────────────
    # (producto_id, sucursal_id, almacen_id) → doc de inventario
    inv_index: dict = {}
    for inv in inventario_raw:
        key = (
            inv.get("producto_id", ""),
            inv.get("sucursal_id", ""),
            inv.get("almacen_id", "default"),
        )
        inv_index[key] = inv

    # -- 8. Construir filas del reporte ----------------------------------------
    print("\n  [*] Construyendo dataset...")
    rows = []

    for producto in productos_raw:
        prod_id         = str(producto["_id"])
        prod_nombre     = producto.get("descripcion", "Sin nombre")
        prod_cod_sis    = producto.get("codigo_sistema", "")
        prod_cod_largo  = str(producto.get("codigo_largo",  "") or "")
        prod_cod_corto  = str(producto.get("codigo_corto",  "") or "")
        prod_costo      = safe_float(producto.get("costo_producto", 0))
        prod_precio     = safe_float(producto.get("precio_venta",   0))
        prod_cat        = categorias_map.get(producto.get("categoria_id", ""), "Sin categoría")
        prod_tipo       = producto.get("tipo_item", "FISICO")

        # Registros de inventario para este producto
        mis_inv = [
            (suc_id, alm_id, doc)
            for (pid, suc_id, alm_id), doc in inv_index.items()
            if pid == prod_id
        ]

        if mis_inv:
            for suc_id, alm_id, inv in mis_inv:
                cantidad         = safe_float(inv.get("cantidad", 0))
                precio_suc_raw   = inv.get("precio_sucursal")
                precio_suc       = safe_float(precio_suc_raw) if precio_suc_raw else None
                precio_efectivo  = precio_suc if (precio_suc and precio_suc > 0) else prod_precio

                if cantidad <= 0:
                    estado = "Sin stock"
                elif cantidad <= 5:
                    estado = "Stock bajo"
                else:
                    estado = "OK"

                rows.append({
                    "Sucursal":             sucursales_map.get(suc_id, suc_id),
                    "Sucursal_ID":          suc_id,
                    "Almacen":              alm_id,
                    "Categoria":            prod_cat,
                    "Producto":             prod_nombre,
                    "Codigo_Sistema":       prod_cod_sis,
                    "Codigo_Corto":         prod_cod_corto,
                    "Codigo_Largo":         prod_cod_largo,
                    "Tipo":                 prod_tipo,
                    "Cantidad":             cantidad,
                    "Costo_Unitario":       prod_costo,
                    "Precio_Base":          prod_precio,
                    "Precio_Sucursal":      precio_suc if precio_suc else "",
                    "Precio_Efectivo":      precio_efectivo,
                    "Valor_Stock":          round(cantidad * prod_costo, 2),
                    "Estado":               estado,
                    "Ultima_Actualizacion": str(inv.get("updated_at", ""))[:19],
                    "Producto_ID":          prod_id,
                    "Inventario_ID":        str(inv.get("_id", "")),
                })
        else:
            # El producto existe en catálogo pero no tiene ningún registro de inventario
            rows.append({
                "Sucursal":             "— Sin asignar —",
                "Sucursal_ID":          "",
                "Almacen":              "—",
                "Categoria":            prod_cat,
                "Producto":             prod_nombre,
                "Codigo_Sistema":       prod_cod_sis,
                "Codigo_Corto":         prod_cod_corto,
                "Codigo_Largo":         prod_cod_largo,
                "Tipo":                 prod_tipo,
                "Cantidad":             0.0,
                "Costo_Unitario":       prod_costo,
                "Precio_Base":          prod_precio,
                "Precio_Sucursal":      "",
                "Precio_Efectivo":      prod_precio,
                "Valor_Stock":          0.0,
                "Estado":               "Sin registro",
                "Ultima_Actualizacion": "",
                "Producto_ID":          prod_id,
                "Inventario_ID":        "",
            })

    print(f"  [OK] {len(rows)} filas en el dataset")

    # ── 9. DataFrames ──────────────────────────────────────────────────────────
    df = pd.DataFrame(rows)

    col_order = [
        "Sucursal", "Almacen", "Categoria", "Producto",
        "Codigo_Sistema", "Codigo_Corto", "Codigo_Largo", "Tipo",
        "Cantidad", "Costo_Unitario", "Precio_Base",
        "Precio_Sucursal", "Precio_Efectivo", "Valor_Stock",
        "Estado", "Ultima_Actualizacion",
        "Producto_ID", "Sucursal_ID", "Inventario_ID",
    ]
    df = df[col_order]

    df_resumen = (
        df.groupby("Sucursal")
        .agg(
            Productos_Distintos=("Producto", "nunique"),
            Total_Unidades=("Cantidad", "sum"),
            Valor_Total_Stock=("Valor_Stock", "sum"),
            Prods_Sin_Stock=("Estado", lambda x: (x == "Sin stock").sum()),
            Prods_Stock_Bajo=("Estado", lambda x: (x == "Stock bajo").sum()),
        )
        .reset_index()
        .sort_values("Total_Unidades", ascending=False)
    )

    prods_con_stock = df[df["Cantidad"] > 0]["Producto_ID"].unique()
    df_sin_stock = (
        df[~df["Producto_ID"].isin(prods_con_stock)][
            ["Codigo_Sistema", "Codigo_Corto", "Producto", "Categoria", "Tipo", "Precio_Efectivo"]
        ]
        .drop_duplicates()
        .sort_values("Categoria")
    )

    # ── 10. Escribir Excel ─────────────────────────────────────────────────────
    print(f"\n  [*] Guardando: {OUTPUT_FILE}")

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="INVENTARIO_COMPLETO", index=False)
        df_resumen.to_excel(writer, sheet_name="RESUMEN_POR_SUCURSAL", index=False)
        if len(df_sin_stock) > 0:
            df_sin_stock.to_excel(writer, sheet_name="PRODUCTOS_SIN_STOCK", index=False)

        wb = writer.book

        # Formatear INVENTARIO_COMPLETO
        ws1 = writer.sheets["INVENTARIO_COMPLETO"]
        format_sheet(ws1, header_color="1E3A5F")

        # Colorear filas según Estado
        try:
            estado_col_idx = col_order.index("Estado") + 1
        except ValueError:
            estado_col_idx = None

        COLOR_MAP = {
            "Sin stock":    "F8D7DA",
            "Stock bajo":   "FFF3CD",
            "Sin registro": "E8E8E8",
            "OK_par":       "D4EDDA",
            "OK_impar":     "FFFFFF",
        }

        if estado_col_idx:
            for row in ws1.iter_rows(min_row=2):
                estado_val = str(row[estado_col_idx - 1].value or "")
                if estado_val in ("Sin stock", "Stock bajo", "Sin registro"):
                    bg = COLOR_MAP[estado_val]
                else:
                    bg = COLOR_MAP["OK_par"] if row[0].row % 2 == 0 else COLOR_MAP["OK_impar"]
                fill = PatternFill("solid", fgColor=bg)
                for cell in row:
                    cell.fill = fill

        format_sheet(writer.sheets["RESUMEN_POR_SUCURSAL"], header_color="2E86AB")

        if "PRODUCTOS_SIN_STOCK" in writer.sheets:
            format_sheet(writer.sheets["PRODUCTOS_SIN_STOCK"], header_color="8B0000")

    # ── 11. Resumen ────────────────────────────────────────────────────────────
    total_uni   = df["Cantidad"].sum()
    total_val   = df["Valor_Stock"].sum()
    sin_stock   = (df["Estado"] == "Sin stock").sum()
    stock_bajo  = (df["Estado"] == "Stock bajo").sum()
    sin_reg     = (df["Estado"] == "Sin registro").sum()

    print(f"\n{'='*62}")
    print(f"  EXPORTACION COMPLETADA")
    print(f"{'='*62}")
    print(f"  Archivo:               {OUTPUT_FILE}")
    print(f"  Productos en catalogo: {len(productos_raw)}")
    print(f"  Registros inventario:  {len(inventario_raw)}")
    print(f"  Sucursales:            {len(sucursales_map)}")
    print()
    print(f"  Totales:")
    print(f"     Unidades en stock:     {total_uni:,.1f}")
    print(f"     Valor del stock:       Bs. {total_val:,.2f}")
    print(f"     Filas sin stock:       {sin_stock}")
    print(f"     Filas stock bajo:      {stock_bajo}")
    print(f"     Sin registro alguno:   {sin_reg}")
    print()
    print(f"  Hojas del Excel:")
    print(f"     1. INVENTARIO_COMPLETO    - Detalle por producto x sucursal x almacen")
    print(f"     2. RESUMEN_POR_SUCURSAL   - Totales agregados por sucursal")
    if len(df_sin_stock) > 0:
        print(f"     3. PRODUCTOS_SIN_STOCK    - {len(df_sin_stock)} productos sin ningun stock")
    print(f"{'='*62}\n")

    await client.close()


if __name__ == "__main__":
    asyncio.run(export_inventario())
